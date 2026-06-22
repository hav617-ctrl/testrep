"""
analyze_dialogue.py

Windows GUI application that analyzes user-AI conversation records in an Excel file
(Sheet 1, columns A=user, B=AI) and identifies up to 3 locations showing
垂直的な深まり (vertical deepening). Applies blue cell shading and evaluation
comments to those cells.
"""

import json
import os
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import anthropic
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment

BLUE_FILL = PatternFill(fill_type="solid", fgColor="ADD8E6")


# ---------------------------------------------------------------------------
# Core analysis logic (no UI dependencies)
# ---------------------------------------------------------------------------

def read_conversation(ws) -> list[dict]:
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        a_val = row[0] if len(row) > 0 else None
        b_val = row[1] if len(row) > 1 else None
        if a_val is None and b_val is None:
            continue
        rows.append({"row": row_idx, "user": a_val or "", "ai": b_val or ""})
    return rows


def build_prompt(rows: list[dict]) -> str:
    lines = []
    for r in rows:
        lines.append(f"行{r['row']} ユーザー: {r['user']}")
        lines.append(f"行{r['row']} AI: {r['ai']}")
        lines.append("")
    conversation_text = "\n".join(lines)

    return f"""以下はユーザーとAIの対話記録です（Excelの行番号付き）。

{conversation_text}

この対話の中で「垂直的な深まり」が生まれている箇所を最大3か所特定してください。
垂直的な深まりとは、会話が表面的な情報交換にとどまらず、より深い洞察・理解・気づき・感情的または知的な発展が生まれている瞬間のことです。

以下のJSON形式で回答してください。他のテキストは含めないでください。

{{
  "selections": [
    {{
      "row": <行番号（整数）>,
      "column": "<'A'または'B'のどちらに深まりが顕れているか>",
      "comment": "<その箇所が垂直的な深まりである理由（100〜200字程度の日本語）>"
    }}
  ]
}}

selectionは最大3件です。最も深まりが顕著な箇所を優先してください。"""


def call_claude(api_key: str, conversation_rows: list[dict]) -> list[dict]:
    client = anthropic.Anthropic(api_key=api_key)
    prompt = build_prompt(conversation_rows)

    response = client.messages.create(
        model="claude-opus-4-8",
        max_tokens=4096,
        thinking={"type": "adaptive"},
        messages=[{"role": "user", "content": prompt}],
    )

    text_content = ""
    for block in response.content:
        if block.type == "text":
            text_content = block.text
            break

    start = text_content.find("{")
    end = text_content.rfind("}") + 1
    if start == -1 or end == 0:
        raise ValueError(f"Claude の応答に JSON が含まれていませんでした:\n{text_content}")

    data = json.loads(text_content[start:end])
    return data.get("selections", [])


def apply_formatting(ws, selections: list[dict]) -> list[str]:
    log_lines = []
    for sel in selections:
        row = sel.get("row")
        col_letter = sel.get("column", "A").upper()
        comment_text = sel.get("comment", "")

        if not row or col_letter not in ("A", "B"):
            continue

        col_idx = 1 if col_letter == "A" else 2
        cell = ws.cell(row=row, column=col_idx)
        cell.fill = BLUE_FILL

        comment = Comment(comment_text, "Claude")
        comment.width = 300
        comment.height = 120
        cell.comment = comment

        log_lines.append(f"  行{row} 列{col_letter}: 青色網掛 + コメント付加")
    return log_lines


# ---------------------------------------------------------------------------
# GUI Application
# ---------------------------------------------------------------------------

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("対話分析 - 垂直的な深まり検出")
        self.resizable(True, True)
        self.minsize(640, 520)

        self._input_path: str = ""
        self._workbook = None
        self._selections: list[dict] = []

        self._build_menu()
        self._build_body()

    # ------------------------------------------------------------------
    # UI construction
    # ------------------------------------------------------------------

    def _build_menu(self):
        menubar = tk.Menu(self)

        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="入力ファイルを開く...", accelerator="Ctrl+O",
                              command=self._open_file)
        file_menu.add_command(label="結果を名前を付けて保存...", accelerator="Ctrl+S",
                              command=self._save_file, state="disabled")
        file_menu.add_separator()
        file_menu.add_command(label="終了", command=self.destroy)
        menubar.add_cascade(label="ファイル", menu=file_menu)

        self.config(menu=menubar)
        self._file_menu = file_menu

        self.bind_all("<Control-o>", lambda _: self._open_file())
        self.bind_all("<Control-s>", lambda _: self._save_file())

    def _build_body(self):
        # ---- API key row ----
        key_frame = tk.Frame(self, padx=8, pady=6)
        key_frame.pack(fill="x")

        tk.Label(key_frame, text="Anthropic API キー:").pack(side="left")
        self._api_key_var = tk.StringVar(value=os.environ.get("ANTHROPIC_API_KEY", ""))
        api_entry = tk.Entry(key_frame, textvariable=self._api_key_var,
                             show="*", width=55)
        api_entry.pack(side="left", padx=6)

        # Toggle show/hide
        self._show_key = False
        self._toggle_btn = tk.Button(key_frame, text="表示",
                                     command=lambda: self._toggle_key_visibility(api_entry))
        self._toggle_btn.pack(side="left")

        # ---- File path row ----
        path_frame = tk.Frame(self, padx=8, pady=4)
        path_frame.pack(fill="x")

        tk.Label(path_frame, text="入力ファイル:     ").pack(side="left")
        self._path_var = tk.StringVar(value="（未選択）")
        tk.Label(path_frame, textvariable=self._path_var,
                 relief="sunken", anchor="w", width=55).pack(side="left", padx=6)
        tk.Button(path_frame, text="参照...", command=self._open_file).pack(side="left")

        # ---- Run / Save buttons ----
        btn_frame = tk.Frame(self, padx=8, pady=6)
        btn_frame.pack(fill="x")
        self._run_btn = tk.Button(btn_frame, text="分析を実行する",
                                  command=self._run_analysis,
                                  state="disabled", width=20)
        self._run_btn.pack(side="left")
        self._save_btn = tk.Button(btn_frame, text="結果を保存...",
                                   command=self._save_file,
                                   state="disabled", width=16)
        self._save_btn.pack(side="left", padx=8)

        # ---- Log area ----
        log_frame = tk.Frame(self, padx=8, pady=4)
        log_frame.pack(fill="both", expand=True)
        tk.Label(log_frame, text="ログ:").pack(anchor="w")
        self._log = scrolledtext.ScrolledText(log_frame, state="disabled",
                                              wrap="word", height=20)
        self._log.pack(fill="both", expand=True)

        # ---- Status bar ----
        self._status_var = tk.StringVar(value="API キーを入力してファイルを選択してください。")
        status_bar = tk.Label(self, textvariable=self._status_var,
                              bd=1, relief="sunken", anchor="w", padx=4)
        status_bar.pack(fill="x", side="bottom")

    def _toggle_key_visibility(self, entry: tk.Entry):
        self._show_key = not self._show_key
        entry.config(show="" if self._show_key else "*")
        self._toggle_btn.config(text="隠す" if self._show_key else "表示")

    # ------------------------------------------------------------------
    # Actions
    # ------------------------------------------------------------------

    def _open_file(self):
        path = filedialog.askopenfilename(
            title="Excelファイルを選択",
            filetypes=[("Excel ファイル", "*.xlsx *.xlsm *.xls"), ("すべてのファイル", "*.*")],
        )
        if not path:
            return

        self._input_path = path
        self._path_var.set(path)
        self._workbook = None
        self._selections = []
        self._file_menu.entryconfig("結果を名前を付けて保存...", state="disabled")
        self._save_btn.config(state="disabled")
        self._run_btn.config(state="normal")
        self._set_status("ファイルを選択しました。「分析を実行する」ボタンを押してください。")
        self._log_clear()
        self._log_write(f"入力ファイル: {path}\n")

    def _run_analysis(self):
        api_key = self._api_key_var.get().strip()
        if not api_key:
            messagebox.showwarning("APIキー未入力",
                                   "Anthropic API キーを入力してください。\n"
                                   "APIキーは https://console.anthropic.com/ で取得できます。")
            return
        if not self._input_path:
            messagebox.showwarning("警告", "先にファイルを選択してください。")
            return

        self._run_btn.config(state="disabled")
        self._save_btn.config(state="disabled")
        self._file_menu.entryconfig("結果を名前を付けて保存...", state="disabled")
        self._set_status("分析中...")
        thread = threading.Thread(target=self._analysis_worker,
                                  args=(api_key,), daemon=True)
        thread.start()

    def _analysis_worker(self, api_key: str):
        try:
            self._log_write("Excelファイルを読み込んでいます...\n")
            wb = load_workbook(self._input_path)
            ws = wb.worksheets[0]

            self._log_write("対話データを解析しています...\n")
            rows = read_conversation(ws)
            if not rows:
                self._on_error("対話データが見つかりませんでした（A列・B列が空です）。")
                return
            self._log_write(f"  {len(rows)} 行の対話を読み込みました\n\n")

            self._log_write("Claude に垂直的な深まりを分析させています...\n"
                            "（APIの応答に数十秒かかる場合があります）\n")
            selections = call_claude(api_key, rows)
            self._log_write(f"  {len(selections)} か所の垂直的な深まりを特定しました\n\n")

            self._log_write("セルに書式を適用しています...\n")
            log_lines = apply_formatting(ws, selections)
            for line in log_lines:
                self._log_write(line + "\n")

            self._log_write("\n--- 分析結果 ---\n")
            for i, sel in enumerate(selections, 1):
                self._log_write(
                    f"\n[{i}] 行{sel.get('row')} 列{sel.get('column')}\n"
                    f"    {sel.get('comment', '')}\n"
                )

            self._workbook = wb
            self._selections = selections

            self.after(0, self._on_success)

        except Exception as exc:
            self.after(0, lambda: self._on_error(str(exc)))

    def _save_file(self):
        if self._workbook is None:
            messagebox.showwarning("警告", "先に分析を実行してください。")
            return

        save_path = filedialog.asksaveasfilename(
            title="結果ファイルを保存",
            defaultextension=".xlsx",
            filetypes=[("Excel ファイル", "*.xlsx"), ("すべてのファイル", "*.*")],
            initialfile="result.xlsx",
        )
        if not save_path:
            return

        try:
            self._workbook.save(save_path)
            self._log_write(f"\n保存しました: {save_path}\n")
            self._set_status(f"保存完了: {save_path}")
            messagebox.showinfo("保存完了", f"ファイルを保存しました:\n{save_path}")
        except Exception as exc:
            messagebox.showerror("保存エラー", str(exc))

    # ------------------------------------------------------------------
    # Callbacks
    # ------------------------------------------------------------------

    def _on_success(self):
        self._run_btn.config(state="normal")
        self._save_btn.config(state="normal")
        self._file_menu.entryconfig("結果を名前を付けて保存...", state="normal")
        self._set_status("分析完了。「結果を保存...」で保存してください。")

    def _on_error(self, message: str):
        self._run_btn.config(state="normal")
        self._set_status("エラーが発生しました。")
        self._log_write(f"\n[エラー] {message}\n")
        messagebox.showerror("エラー", message)

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _log_write(self, text: str):
        def _do():
            self._log.config(state="normal")
            self._log.insert("end", text)
            self._log.see("end")
            self._log.config(state="disabled")
        self.after(0, _do)

    def _log_clear(self):
        self._log.config(state="normal")
        self._log.delete("1.0", "end")
        self._log.config(state="disabled")

    def _set_status(self, text: str):
        self._status_var.set(text)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
