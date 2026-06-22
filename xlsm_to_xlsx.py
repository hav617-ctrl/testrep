"""
複数の .xlsm ファイルを読み込み、空でない最初のシートだけを
同名の .xlsx ファイルとして保存するスクリプト。

使い方:
  python xlsm_to_xlsx.py                  # カレントディレクトリの全 .xlsm を処理
  python xlsm_to_xlsx.py file1.xlsm ...   # 指定ファイルを処理
  python xlsm_to_xlsx.py --dir /path/to   # 指定ディレクトリの全 .xlsm を処理
  python xlsm_to_xlsx.py --out /path/to   # 出力先ディレクトリを指定（省略時は入力と同じ場所）
"""

import argparse
import sys
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl import Workbook


def is_sheet_empty(sheet) -> bool:
    """シートが空（値を持つセルが1つもない）かどうかを返す。"""
    return all(cell.value is None for row in sheet.iter_rows() for cell in row)


def find_first_nonempty_sheet(wb):
    """ワークブック内で空でない最初のシートを返す。見つからなければ None。"""
    for name in wb.sheetnames:
        ws = wb[name]
        if not is_sheet_empty(ws):
            return ws
    return None


def copy_sheet_to_new_workbook(src_ws) -> Workbook:
    """A・B列のみ新しいワークブックへコピーして返す。列幅は内容に合わせて自動調整。"""
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = src_ws.title

    max_width = {1: 0, 2: 0}  # 列ごとの最大文字数

    # A列(1)・B列(2)のみコピー
    for row in src_ws.iter_rows():
        for cell in row:
            if cell.column not in (1, 2):
                continue
            new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
            # 列幅計算用に文字数を記録
            length = len(str(cell.value)) if cell.value is not None else 0
            max_width[cell.column] = max(max_width[cell.column], length)

    # 列幅を内容に合わせて設定（最小8、最大60）
    for col_num, width in max_width.items():
        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(col_num)
        new_ws.column_dimensions[col_letter].width = max(8, min(width + 4, 60))

    # 行高をコピー
    for row_num, row_dim in src_ws.row_dimensions.items():
        new_ws.row_dimensions[row_num].height = row_dim.height
        new_ws.row_dimensions[row_num].hidden = row_dim.hidden

    return new_wb


def process_file(xlsm_path: Path, out_dir: Path) -> None:
    print(f"処理中: {xlsm_path}")
    try:
        wb = openpyxl.load_workbook(xlsm_path, keep_vba=False, data_only=True)
    except Exception as e:
        print(f"  [エラー] 読み込み失敗: {e}", file=sys.stderr)
        return

    ws = find_first_nonempty_sheet(wb)
    if ws is None:
        print(f"  [スキップ] 空でないシートが見つかりません")
        return

    print(f"  使用シート: {ws.title!r}")
    new_wb = copy_sheet_to_new_workbook(ws)

    out_path = out_dir / (xlsm_path.stem + ".xlsx")
    try:
        new_wb.save(out_path)
        print(f"  保存完了: {out_path}")
    except Exception as e:
        print(f"  [エラー] 保存失敗: {e}", file=sys.stderr)


def main():
    parser = argparse.ArgumentParser(description=".xlsm → .xlsx 変換スクリプト")
    parser.add_argument("files", nargs="*", help="処理する .xlsm ファイル")
    parser.add_argument("--dir", default=".", help="入力ディレクトリ（files 未指定時）")
    parser.add_argument("--out", default=None, help="出力ディレクトリ（省略時は入力ファイルと同じ場所）")
    args = parser.parse_args()

    if args.files:
        targets = []
        for f in args.files:
            matched = sorted(Path(".").glob(f)) if ("*" in f or "?" in f) else [Path(f)]
            targets.extend(matched)
    else:
        targets = sorted(Path(args.dir).glob("*.xlsm"))

    if not targets:
        print("処理対象の .xlsm ファイルが見つかりませんでした。")
        return

    for path in targets:
        if not path.exists():
            print(f"[エラー] ファイルが存在しません: {path}", file=sys.stderr)
            continue
        out_dir = Path(args.out) if args.out else path.parent
        out_dir.mkdir(parents=True, exist_ok=True)
        process_file(path, out_dir)


if __name__ == "__main__":
    main()
