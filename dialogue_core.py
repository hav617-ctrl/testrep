"""
dialogue_core.py

Shared core logic used by both the Windows and Android versions.
"""

import json
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment

BLUE_FILL = PatternFill(fill_type="solid", fgColor="ADD8E6")


def read_conversation(ws) -> list[dict]:
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        a_val = row[0] if len(row) > 0 else None
        b_val = row[1] if len(row) > 1 else None
        if a_val is None and b_val is None:
            continue
        rows.append({"row": row_idx, "user": a_val or "", "ai": b_val or ""})
    return rows


def build_prompt(rows: list[dict]) -> str:
    lines = []
    for r in rows:
        lines.append(f"行{r['row']} ユーザー: {r['user']}")
        lines.append(f"行{r['row']} AI: {r['ai']}")
        lines.append("")
    conversation_text = "\n".join(lines)

    return f"""以下はユーザーとAIの対話記録です（Excelの行番号付き）。

{conversation_text}

この対話の中で「垂直的な深まり」が生まれている箇所を最大3か所特定してください。
垂直的な深まりとは、会話が表面的な情報交換にとどまらず、より深い洞察・理解・気づき・感情的または知的な発展が生まれている瞬間のことです。

以下のJSON形式で回答してください。他のテキストは含めないでください。

{{
  "selections": [
    {{
      "row": <行番号（整数）>,
      "column": "<'A'または'B'のどちらに深まりが顕れているか>",
      "comment": "<その箇所が垂直的な深まりである理由（100〜200字程度の日本語）>"
    }}
  ]
}}

selectionは最大3件です。最も深まりが顕著な箇所を優先してください。"""


def apply_formatting(ws, selections: list[dict]) -> list[str]:
    log_lines = []
    for sel in selections:
        row = sel.get("row")
        col_letter = sel.get("column", "A").upper()
        comment_text = sel.get("comment", "")

        if not row or col_letter not in ("A", "B"):
            continue

        col_idx = 1 if col_letter == "A" else 2
        cell = ws.cell(row=row, column=col_idx)
        cell.fill = BLUE_FILL

        comment = Comment(comment_text, "Claude")
        comment.width = 300
        comment.height = 120
        cell.comment = comment

        log_lines.append(f"  行{row} 列{col_letter}: 青色網掛 + コメント付加")
    return log_lines
